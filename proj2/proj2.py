import os
import pyrebase
import streamlit as st
import pandas as pd
from itertools import repeat
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl import workbook, load_workbook
from openpyxl.styles.borders import Border, Side
from streamlit_login_auth_ui.widgets import __login__

start_time = datetime.now()


thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))


def find_octant(a, b, c):
    if (a > 0 and b > 0 and c > 0):
        return 1
    elif (a > 0 and b > 0 and c < 0):
        return -1
    elif (a < 0 and b > 0 and c > 0):
        return 2
    elif (a < 0 and b > 0 and c < 0):
        return -2
    elif (a < 0 and b < 0 and c > 0):
        return 3
    elif (a < 0 and b < 0 and c < 0):
        return -3
    elif (a > 0 and b < 0 and c > 0):
        return 4
    elif (a > 0 and b < 0 and c < 0):
        return -4


def process_file(s, mod, name):										# Function to process individual file
    in_file = 'input/'+name											# Path to file
    df = pd.read_excel(s)										# Reading file in dataframe
    wb = Workbook()												# Starting a new workbook
    myworkbook = wb.active
    lst = ['T', 'U', 'V', 'W', 'U Avg', 'V Avg', 'W Avg',
           r"U'=U-U avg", r"V'=V-V avg", r"W'=W-W avg", 'Octant']
    for i in range(11):											# Writing header of file
        myworkbook.cell(row=2, column=i+1).value = lst[i]
    octant = []
    # Finding average of u,v and w
    u_avg = df['U'].mean()
    v_avg = df['V'].mean()
    w_avg = df['W'].mean()

    myworkbook.cell(row=1, column=14).value = 'Overall Octant Count'
    myworkbook.cell(row=1, column=45).value = 'Longest Subsequence Length'
    myworkbook.cell(
        row=1, column=49).value = 'Longest Subsquence Length with Range'

    myworkbook.cell(row=3, column=5).value = round(u_avg, 3)
    myworkbook.cell(row=3, column=6).value = round(v_avg, 3)
    myworkbook.cell(row=3, column=7).value = round(w_avg, 3)

    for i in df.index:
        myworkbook.cell(row=i+3, column=1).value = df['T'][i]
        myworkbook.cell(row=i+3, column=2).value = df['U'][i]
        myworkbook.cell(row=i+3, column=3).value = df['V'][i]
        myworkbook.cell(row=i+3, column=4).value = df['W'][i]
        myworkbook.cell(
            row=i+3, column=8).value = round(df['U'][i]-u_avg, 3)
        myworkbook.cell(
            row=i+3, column=9).value = round(df['V'][i]-v_avg, 3)
        myworkbook.cell(
            row=i+3, column=10).value = round(df['W'][i]-w_avg, 3)
        myworkbook.cell(row=i+3, column=11).value = find_octant(
            df['U'][i]-u_avg, df['V'][i]-v_avg, df['W'][i]-w_avg)
        octant.append(find_octant(
            df['U'][i]-u_avg, df['V'][i]-v_avg, df['W'][i]-w_avg))

    def octant_range_names(mod=5000):
        octant_name_id_mapping = {"1": "Internal outward interaction", "-1": "External outward interaction", "2": "External Ejection",
                                  "-2": "Internal Ejection", "3": "External inward interaction", "-3": "Internal inward interaction", "4": "Internal sweep", "-4": "External sweep"}
        # creating dictionary for mapping
        dic = {}
        # Creating dictionary with opposite key value pair than 'dic'
        my_dic = {}

        # dic[1]=0,dic[-1]=-1,...
        for i in range(0, 4):
            # my_dic[0]=1,my_dic[1]=-1,...
            dic[i+1] = 2*i+1-1
            dic[-(i+1)] = 2*(i+1)-1
            my_dic[2*i+1-1] = i+1
            my_dic[2*(i+1)-1] = -(i+1)

        # Function to find the rank list from count values of all octants
        def find_rank_of_list(lst):
            temp_lst = lst.copy()
            temp_lst.sort(reverse=True)
            res = []

            for i in lst:
                for j in range(0, 8):
                    if (i == temp_lst[j]):
                        res.append(j+1)
                        break
            # Returning the ranked list
            return res

        # Finding the octant which has rank 1 in the given rank list
        def find_1st_rank(lst):
            for i in range(8):
                if (lst[i] == 1):
                    return my_dic[i]

        # Finding the count of rank 1 in the rank 1 mod values of octant x
        def count_rank1(lst, x):
            sum = 0
            for i in lst:
                if (x == i):
                    sum += 1
            return sum                                                  # Return the count

        # Matrix to store rank list for different mod values
        my_matr = []
        # List to store the octants which have rank 1 in different mod ranges and overall
        rank1_list = []
        myworkbook = wb.active
        # Putting the string 'User Input' at its specified place
        myworkbook['M4'] = 'Mod '+str(mod)

        # 2-d matrix for storing octants within ranges
        matrix = []
        # Creating a list for storing elements of 9 columns
        count = [0]*9

        # Storing header list in 'count' list
        count[0] = 'Octant ID'

        for i in range(0, 4):
            count[2*i+1] = (i+1)
            count[2*(i+1)] = -(i+1)
        # Appending header list in matrix
        matrix.append(count)
        # Writing header list in worksheet
        for i in range(13, 22):
            myworkbook.cell(row=3, column=i+1).value = count[i-13]
            myworkbook.cell(row=3, column=i+1).border = thin_border
            if (i > 13):
                myworkbook.cell(row=3, column=i +
                                9).value = 'Rank Octant '+str(count[i-13])
                myworkbook.cell(row=3, column=i+9).border = thin_border
        myworkbook.cell(row=3, column=31).value = 'Rank1 Octant ID'
        myworkbook.cell(row=3, column=32).value = 'Rank1 Octant Name'
        myworkbook.cell(row=3, column=31).border = thin_border
        myworkbook.cell(row=3, column=32).border = thin_border
        # Resetting values in list 'count'
        count = [0]*9

        # Finding total count of values in different octants
        for i in octant:
            if (i == 1):
                count[1] = count[1]+1
            elif (i == -1):
                count[2] = count[2]+1
            elif (i == 2):
                count[3] = count[3]+1
            elif (i == -2):
                count[4] = count[4]+1
            elif (i == 3):
                count[5] = count[5]+1
            elif (i == -3):
                count[6] = count[6]+1
            elif (i == 4):
                count[7] = count[7]+1
            elif (i == -4):
                count[8] = count[8]+1
        yellow = "00FFFF00"
        # Creating overall count row
        count[0] = 'Overall Count'
        matrix.append(count)
        # Writing overall count in worksheet
        for i in range(13, 22):
            myworkbook.cell(row=4, column=i+1).value = count[i-13]
            myworkbook.cell(row=4, column=i+1).border = thin_border
        # Removing the header from list
        count.pop(0)
        # Find the rank list
        rank = find_rank_of_list(count)
        # Finding the rank 1 octant and appending in rank1_list
        rank1_list.append(find_1st_rank(rank))
        # Appending rank list in the matrix
        my_matr.append(rank)
        # Writing overall count in worksheet
        for i in range(8):
            myworkbook.cell(row=4, column=23+i).value = my_matr[0][i]
            myworkbook.cell(row=4, column=23+i).border = thin_border
            if (my_matr[0][i] == 1):
                myworkbook.cell(row=4, column=23+i).fill = PatternFill(
                    start_color=yellow, end_color=yellow, fill_type="solid")
        myworkbook.cell(row=4, column=31).value = rank1_list[0]
        myworkbook.cell(
            row=4, column=32).value = octant_name_id_mapping[str(rank1_list[0])]
        myworkbook.cell(row=4, column=31).border = thin_border
        myworkbook.cell(row=4, column=32).border = thin_border

        # Finding the number of points given in the input
        n = len(octant)
        # Resetting the values in the list 'count'
        count = [0]*9
        # Variable to keep track of the index of data we are on
        k = 0
        # Variable to keep track of row in worksheet
        j = 4
        # Counting number of values in different octants in mod range
        for i in octant:
            if (i == 1):
                count[1] = count[1]+1
            elif (i == -1):
                count[2] = count[2]+1
            elif (i == 2):
                count[3] = count[3]+1
            elif (i == -2):
                count[4] = count[4]+1
            elif (i == 3):
                count[5] = count[5]+1
            elif (i == -3):
                count[6] = count[6]+1
            elif (i == 4):
                count[7] = count[7]+1
            elif (i == -4):
                count[8] = count[8]+1
            # Incrementing the index tracking variable
            k = k+1
            # Processing the mod values in the range and storing them in the list 'count'
            if (k % mod == 1):
                count[0] = str(k-1)+'-'
            elif (k % mod == 0 or k == n):
                # Here count[0]-> represents the range and further elements of count represents the count in different octants
                count[0] = count[0]+str(k-1)
                # Writing the mod count of octant in worksheet
                for i in range(13, 22):
                    myworkbook.cell(row=j+1, column=i +
                                    1).value = count[i-13]
                    myworkbook.cell(row=j+1, column=i +
                                    1).border = thin_border
                # Removing the header from list
                count.pop(0)
                # Find the rank list
                rank = find_rank_of_list(count)
                # Finding the rank 1 octant and appending in rank1_list
                rank1_list.append(find_1st_rank(rank))
                # Appending rank list in the matrix
                my_matr.append(rank)

                # Writing the columns of rank, rank1 and octant_name in the worksheet
                for i in range(8):
                    myworkbook.cell(row=j+1, column=23 +
                                    i).value = my_matr[j-3][i]
                    myworkbook.cell(row=j+1, column=23 +
                                    i).border = thin_border
                    if (my_matr[j-3][i] == 1):
                        myworkbook.cell(row=j+1, column=23+i).fill = PatternFill(
                            start_color=yellow, end_color=yellow, fill_type="solid")

                myworkbook.cell(row=j+1, column=31).value = rank1_list[j-3]
                myworkbook.cell(
                    row=j+1, column=32).value = octant_name_id_mapping[str(rank1_list[j-3])]
                myworkbook.cell(row=j+1, column=31).border = thin_border
                myworkbook.cell(row=j+1, column=31).border = thin_border

                j = j+1                                                   # Incrementing row
                matrix.append(count)
                # Resetting count of values in different octants
                count = [0]*9

        # Removing the overall rank1 octant
        rank1_list.pop(0)

        # Writing the header of table of count of rank1 mod values
        myworkbook.cell(row=12, column=29).value = 'Octant ID'
        myworkbook.cell(row=12, column=30).value = 'Octant Name'
        myworkbook.cell(
            row=12, column=31).value = 'Count of Rank 1 Mod Values'
        myworkbook.cell(row=12, column=29).border = thin_border
        myworkbook.cell(row=12, column=30).border = thin_border
        myworkbook.cell(row=12, column=31).border = thin_border

        # Writing the table of count of rank1 mod values
        for i in range(8):
            myworkbook.cell(row=13+i, column=29).value = my_dic[i]
            myworkbook.cell(
                row=13+i, column=30).value = octant_name_id_mapping[str(my_dic[i])]
            myworkbook.cell(
                row=13+i, column=31).value = count_rank1(rank1_list, my_dic[i])
            myworkbook.cell(row=13+i, column=29).border = thin_border
            myworkbook.cell(row=13+i, column=30).border = thin_border
            myworkbook.cell(row=13+i, column=31).border = thin_border

    def octant_longest_subsequence_count_with_range():

        r = ['Count', 'Longest Subsequence Length',
             'Count']                # Header list
        # Writing header of table to worksheet
        for i in range(3):
            myworkbook.cell(row=3, column=45+i).value = r[i]
            myworkbook.cell(row=3, column=45+i).border = thin_border

        octants = []
        # Writing octants on leftmost column of the table
        for i in range(2, 10, 2):
            myworkbook.cell(row=i+2, column=45).value = i//2
            octants.append(i//2)
            myworkbook.cell(row=i+3, column=45).value = -(i//2)
            octants.append(-i//2)
            myworkbook.cell(row=i+2, column=45).border = thin_border
            myworkbook.cell(row=i+3, column=45).border = thin_border

        # creating dictionary for mapping
        dic = {}
        for i in range(0, 4):
            dic[i+1] = 2*i+1-1
            dic[-(i+1)] = 2*(i+1)-1

        # List for storing number of longest subsequence
        count = [0]*8
        # List for storing length of longest subsequence
        longest_length = [0]*8
        prev = octant[0]
        # Length of current octant
        l = 1
        n = len(octant)
        # Temporary variable to store range
        temp = [0]
        # Empty list of list to store ranges for different octants
        ranges = [[] for x in repeat(None, 8)]

        # Loop for finding number and length of longest subsequence
        for i in range(1, n+1):
            # IF last is reached process the whole
            if (i == n):
                if (longest_length[dic[prev]] < l):
                    longest_length[dic[prev]] = l
                    count[dic[prev]] = 1
                    # Writing ending range in temp
                    temp.append(df['T'][i-1])
                    # Clearing range for that octant because current longest length is small
                    ranges[dic[prev]].clear()
                    # Writing longest range for specific octant
                    ranges[dic[prev]].append(temp)
                elif (longest_length[dic[prev]] == l):
                    count[dic[prev]] += 1
                    temp.append(df['T'][i-1])
                    # Appending more ranges to the octant
                    ranges[dic[prev]].append(temp)
            # If prev and current values are same, increase current length by 1
            elif (prev == octant[i]):
                l += 1
            # Else process the previous octant values and start with new octant
            else:
                if (longest_length[dic[prev]] < l):
                    longest_length[dic[prev]] = l
                    count[dic[prev]] = 1
                    # Clearing range for that octant because current longest length is small
                    ranges[dic[prev]].clear()
                    # Writing ending range in temp
                    temp.append(df['T'][i-1])
                    # Writing longest range for specific octant
                    ranges[dic[prev]].append(temp)
                elif (longest_length[dic[prev]] == l):
                    count[dic[prev]] += 1
                    temp.append(df['T'][i-1])
                    # Appending more ranges to the octant
                    ranges[dic[prev]].append(temp)
                # Writing starting of range in temp variable
                temp = [df['T'][i]]
                l = 1
                # Updating previous octant for next octant
                prev = octant[i]

        # Writing the number and length of longest subsequence in table
        for i in range(2, 10):
            myworkbook.cell(row=i+2, column=46).value = longest_length[i-2]
            myworkbook.cell(row=i+2, column=47).value = count[i-2]
            myworkbook.cell(row=i+2, column=46).border = thin_border
            myworkbook.cell(row=i+2, column=47).border = thin_border
        # Variable to keep track of row in worksheet
        k = 2
        # Writing heading of table
        myworkbook.cell(row=k+1, column=49).value = 'Octant ###'
        myworkbook.cell(
            row=k+1, column=50).value = 'Longest Subsequence Length'
        myworkbook.cell(row=k+1, column=51).value = 'Count'
        myworkbook.cell(row=k+1, column=49).border = thin_border
        myworkbook.cell(row=k+1, column=50).border = thin_border
        myworkbook.cell(row=k+1, column=51).border = thin_border

        k += 2
        for i in range(8):
            # Writing contents of table-1 for each octant
            myworkbook.cell(row=k, column=49).value = octants[i]
            myworkbook.cell(row=k, column=50).value = longest_length[i]
            myworkbook.cell(row=k, column=51).value = count[i]
            # Writing header of ranges in worksheet
            myworkbook.cell(row=k+1, column=49).value = 'Time'
            myworkbook.cell(row=k+1, column=50).value = 'From'
            myworkbook.cell(row=k+1, column=51).value = 'To'
            # Adding border to cells
            myworkbook.cell(row=k, column=49).border = thin_border
            myworkbook.cell(row=k, column=50).border = thin_border
            myworkbook.cell(row=k, column=51).border = thin_border
            myworkbook.cell(row=k+1, column=49).border = thin_border
            myworkbook.cell(row=k+1, column=50).border = thin_border
            myworkbook.cell(row=k+1, column=51).border = thin_border
            x = ranges[i]
            k += 2
            for j in x:
                # Writing ranges in worksheet
                myworkbook.cell(row=k, column=50).value = j[0]
                myworkbook.cell(row=k, column=51).value = j[1]
                # Adding border to cells
                myworkbook.cell(row=k, column=49).border = thin_border
                myworkbook.cell(row=k, column=50).border = thin_border
                myworkbook.cell(row=k, column=51).border = thin_border
                k += 1

    def octant_transition_count(mod=5000):
        j = 1
        n = len(octant)
        # Writing overall transition count in worksheet
        myworkbook.cell(
            row=j, column=35).value = 'Overall Transition Count'
        myworkbook.cell(row=j+3, column=34).value = 'From'
        myworkbook.cell(row=j+1, column=36).value = 'To'
        j += 2

        # Creating 9*9 matrix for storing transition count values
        matrix = [[0]*9 for i in range(9)]

        # Storing header row and header column in the matrix
        for i in range(0, 4):
            matrix[0][2*i+1] = (i+1)
            matrix[0][2*(i+1)] = -(i+1)
        for i in range(0, 9):
            matrix[i][0] = matrix[0][i]
        matrix[0][0] = 'Octant #'

        # creating dictionary for mapping
        dic = {}
        for i in range(0, 4):
            dic[i+1] = 2*i+1
            dic[-(i+1)] = 2*(i+1)

        # Finding row and column of matrix from transition values
        def find_row_col(x, y):
            lst = [dic[x], dic[y]]
            return lst

        def find_max_ele(lst):
            temp = lst.copy()
            temp.pop(0)
            large = 0
            for i in temp:
                if (large < i):
                    large = i
            return large

        prev = octant[0]
        # Filling overall transition matrix
        for i in range(1, n):
            # lst[0]-> row and lst[1]->column of overall transition matrix
            lst = find_row_col(prev, octant[i])
            matrix[lst[0]][lst[1]] += 1
            prev = octant[i]
        yellow = "00FFFF00"
        # Writing the overall transition matrix in worksheet
        for i in range(0, 9):
            temp_lst = matrix[i]
            large = find_max_ele(temp_lst)
            for k in range(13, 22):
                myworkbook.cell(row=j+i, column=k +
                                22).value = matrix[i][k-13]
                myworkbook.cell(row=j+i, column=k+22).border = thin_border
                if (i > 0 and matrix[i][k-13] == large):
                    myworkbook.cell(row=j+i, column=k+22).fill = PatternFill(
                        start_color=yellow, end_color=yellow, fill_type="solid")
                if (i != 0 and k != 13):
                    matrix[i][k-13] = 0

        # temp-> No. of mod transition tables
        temp = n//mod+1
        j += 1
        # One iteration for each mod transition table
        for t in range(0, temp):
            j += 11
            name = ''
            # Writing Table name in worksheet
            myworkbook.cell(
                row=j, column=35).value = 'Mod Transition Count'
            myworkbook.cell(row=j+3, column=34).value = 'From'
            myworkbook.cell(row=j+1, column=36).value = 'To'
            name = str(t*mod)+'-'
            if ((t+1)*mod-1 > n-1):
                name += str(n-1)
            else:
                name += str((t+1)*mod-1)
            myworkbook.cell(row=j+1, column=35).value = name
            j += 2

            # Incrementing matrix cell corresponding to transition values
            for i in range(t*mod, min(n-1, (t+1)*mod)):
                lst = find_row_col(octant[i], octant[i+1])
                matrix[lst[0]][lst[1]] += 1

            # Writing the transition mod matrix in worksheet
            for i in range(0, 9):
                temp_lst = matrix[i]
                if (i > 0):
                    large = find_max_ele(temp_lst)
                for k in range(13, 22):
                    myworkbook.cell(row=j+i, column=k +
                                    22).value = matrix[i][k-13]
                    myworkbook.cell(row=j+i, column=k +
                                    22).border = thin_border
                    if (i > 0 and matrix[i][k-13] == large):
                        myworkbook.cell(row=j+i, column=k+22).fill = PatternFill(
                            start_color=yellow, end_color=yellow, fill_type="solid")
                    if (i != 0 and k != 13):
                        # Resetting matrix for next mod iteration
                        matrix[i][k-13] = 0

    octant_transition_count(mod=5000)
    octant_range_names(5000)
    octant_longest_subsequence_count_with_range()
    # s = s[:-5]
    file_name = 'output/'+name+' cm_vel_octant_analysis_mod_'+str(mod)+'.xlsx'
    wb.save(file_name)
    st.write(file_name)


def octant_analysis(mod, file, multiple_file_bool):
    path = os.getcwd()
    if not os.path.exists(os.path.join(path, "output")):
        os.mkdir('output')
        print("Directory ", 'output',  " created")
    else:
        print("Directory ", 'output',  " already exists")
    # input_files = os.listdir('input')
    if multiple_file_bool:
        for i in range(len(file)):
            process_file(file[i], mod, file[i].name)
    else:
        process_file(file, mod, file.name)


# users = db.fetch_All_user()

# usernames = [user["key"] for user in users]
# names = [user["name"] for user in users]
# hashed_passwords = [user["password"] for user in users]

# authenticator = stauth.Authenticate(
#     names, usernames, hashed_passwords, "proj2", "!#@R!2", cookie_expiry_days=30)

# name, authentication_status, username = authenticator.login("Login", "main")


# --------------------------------------Configuration for firebase----------------------------
# firebaseConfig = {
#     'apiKey': "AIzaSyDjwMZSSyKkvoUmN980-5PUM48FhMmWIsk",
#     'authDomain': "cs384proj.firebaseapp.com",
#     'projectId': "cs384proj",
#     'storageBucket': "cs384proj.appspot.com",
#     'messagingSenderId': "533524803259",
#     'appId': "1:533524803259:web:8d87b0e216e547f299039e",
#     'measurementId': "G-VK31SVCY7G",
#     'databaseURL': 'https://cs384proj-default-rtdb.europe-west1.firebasedatabase.app/'
# }


# -------------------------------------- firebase authentication init----------------------------
# firebase = pyrebase.initialize_app(firebaseConfig)
# auth = firebase.auth()

# # databse
# db = firebase.database()


# -------------------------------------------Login/Sign up Page------------------------------------------------

# @st.cache(suppress_st_warning=True)
# def setup_login_signup_page():
#     holder_choice = st.empty()
#     choice = holder_choice.selectbox('login or signup', ['Login', 'Sign up'])

#     if choice == 'Login':

#         holder_email = st.empty()
#         email = holder_email.text_input('Please enter your email address')

#         holder_password = st.empty()
#         password = holder_password.text_input(
#             'Please Enter your password', type='password')

#         holder_button = st.empty()
#         submit = holder_button.button('Login', 'login')
#         # if submit and email == None:
#         #     st.warning('Pls Enter your email')
#         # elif submit and password == None:
#         #     st.warning('Pls Enter Password')
#         if submit:
#             if email == None:
#                 st.warning('Pls Enter your email')
#             if password == None:
#                 st.warning('Pls Enter Password')
#             else:
#                 user = auth.sign_in_with_email_and_password(email, password)
#                 st.balloons()
#                 st.info('Logined successfully')
#                 holder_button.empty()
#                 holder_password.empty()
#                 holder_choice.empty()
#                 holder_email.empty()
#     else:
#         holder_email = st.empty()
#         email = holder_email.text_input('Please enter your email address')

#         holder_password = st.empty()
#         password = holder_password.text_input(
#             'Please Enter your password', type='password')

#         holder_submit = st.empty()
#         submit = holder_submit.button('Create my account', 'signup')

#         if submit:
#             if email == None:
#                 st.warning('Pls Enter your email')
#             if password == None:
#                 st.warning('Pls Enter Password')
#             else:
#                 user = auth.create_user_with_email_and_password(
#                     email, password)
#                 st.balloons()
#                 st.info('Account Created successfully')
#                 holder_submit.empty()
#                 holder_email.empty()
#                 holder_password.empty()
#                 holder_choice.empty()


# ------------------------------setting up login/signup form--------------------
# setup_login_signup_page()
# t1 = threading.Thread(target=setup_login_signup_page)
# t1.start()
# t1.join()
# @st.experimental_singleton
# def choice():
#     st.selectbox("lau", ['hehe', 'khekhe'])


# choice()

# if st.button('clear all'):
#     st.experimental_singleton.clear()

__login__obj = __login__(auth_token="pk_prod_6SW5WQ1TNX4YBAHQJZPHJA95AJTG",
                         company_name="cs384",
                         width=200, height=250,
                         logout_button_name='Logout', hide_menu_bool=False,
                         hide_footer_bool=False,
                         lottie_url='https://lottie.host/869c634e-782a-4759-af25-be04b82ba59f/wmK9UTLckB.json')

LOGGED_IN = __login__obj.build_login_ui()

# if LOGGED_IN == True:

#     st.markown("Your Streamlit Application Begins here!")

if LOGGED_IN != True:
    # st.error("Username or password is incorrect")
    pass
else:

    # ----------------------sidebar-------------------
    # defininf the radio button
    radio_box_list = ['Upload Single File', 'Upload Multiple Files']
    radio_box = st.sidebar.radio(
        'Uploading Options', radio_box_list, horizontal=True)

    # global mod

    # different options to upload files/files according to radio box selected respectively.
    if radio_box == 'Upload Single File':
        upload_single_file = st.sidebar.file_uploader(
            "Choose a file", type=['xlsx'], accept_multiple_files=False)
        mod = st.number_input('Enter mod value', value=5000)
        s_button = st.button('Compute', key='compute single file')
        if s_button and upload_single_file == None:
            st.warning('Please upload a file first')
        elif s_button:
            octant_analysis(mod, upload_single_file, multiple_file_bool=False)
        # else:
        #     pass

    elif radio_box == 'Upload Multiple Files':
        upload_multiple_file = st.sidebar.file_uploader(
            "Choose a Folder", type=['xlsx'], accept_multiple_files=True)
        mod = st.number_input('Enter mod value', value=5000)
        m_button = st.button('Compute', key='compute single file')
        if m_button and upload_multiple_file == None:
            st.warning('Please upload a file first')
        elif m_button:
            octant_analysis(mod, upload_multiple_file, multiple_file_bool=True)
        # else:
        #     pass


# This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
